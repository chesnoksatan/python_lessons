#todo:
#   Для задачи task12.py "Морской бой", написать Save game. Пользователь может прервать игру и сохраниться, затем продолжить либо выйти.
#   Предусмотереть возможность восстановить игру из заранее сохраненного состояния. Сохранение произвести в файл по имени.

#todo:
# Тестовая система выгружает файл в следующем формате:
# {
#     1: 'Пупкин',
#     .......,
#     12. 'Васичкин'
# }
#
# id1 - дисятичное целое равное номеру задачи
# [{ 'id': 1 , 'fail': [ id3, id5], 'pass': [id1, id2, id4]}, { 'id': 12 , fail: [ id1, id2, id5], 'pass': [id3, id4]} ,.. ]
# Ключи словаря:
#  id - идентификатор студента прошедшего тест
#  pass - количество пройденных задач
#  fail - количесто не пройденных задач
#
#
#
# Необходимо сформировать Excel файл для последующей загрузки в  Spreadsheets Google где в качестве шапки - фамилии, а в качестве
# первого номера - задачи. Поля при этом должны быть заполнены + или - в зависимости от результата проверки.
#
# Материал:
# https://realpython.com/openpyxl-excel-spreadsheets-python/

import json
from openpyxl import Workbook
from argparse import ArgumentParser


def read_file(path: str):
    try:
        with open(path, "r") as file:
            return json.load(file)
    except Exception:
        return []
    
def write_file(workbook: Workbook, path: str | None) -> None:
    if path is None:
        workbook.save(filename="exam.xlsx")
    else:
        workbook.save(filename=path)

def create_sheet(workbook: Workbook, exam: list, students: dict) -> None:
    sheet = workbook.active

    students_count = len(students)

    print(students)

    def print_rows():
        for row in sheet.iter_rows(values_only=True):
            print(row)

    for index, id in enumerate(students):
        print(index, id)
        sheet.insert_cols(idx=index+1)
        sheet.cell(1, index+1).value = students[id]
    
    print_rows()

    # for row in sheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=students_count):
    #     print(row)

    # sheet["B1"] = students["key"]

    # for student in exam:
    #     student_id = student[id]


version = "0.0.1"
parser = ArgumentParser(prog="Exam json to xlsx")
parser.add_argument("-v", "--version", action="version", version=version)
parser.add_argument("students", help="Принимает файл json со студентами")
parser.add_argument("input", help="Принимает файл json c результатами экзамена")
parser.add_argument('--output', help="Выходной файл")

if __name__ == "__main__":
    workbook = Workbook()
    
    args = parser.parse_args()

    exam = read_file(args.input)
    students = read_file(args.students)

    print(students, exam)

    # create_sheet(workbook, exam, students)

    # sheet["B1"] = "1"
    # sheet["A2"] = "id1"
    # sheet["B2"] = "+"
    
    # write_file(workbook, args.output)